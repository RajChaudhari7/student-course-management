from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


def export_to_excel(filename, report_title, headers, rows):

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Report Title
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(headers)
    )

    title_cell = ws["A1"]
    title_cell.value = report_title
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Generated Time
    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=len(headers)
    )

    info = ws["A2"]
    info.value = (
        "Generated on: "
        + datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    )

    info.font = Font(italic=True)

    # Header Row
    header_fill = PatternFill(
        start_color="0D6EFD",
        end_color="0D6EFD",
        fill_type="solid"
    )

    row_number = 4

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=row_number, column=col)

        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data Rows
    for row in rows:
        ws.append(row)

    # Auto Width
    for column in ws.columns:

        max_length = 0
        letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except:
                pass

        ws.column_dimensions[letter].width = max_length + 4

    wb.save(filename)

    return filename